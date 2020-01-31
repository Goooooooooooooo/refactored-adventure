from PIL import Image
import openpyxl
from openpyxl.styles import fills
from openpyxl.utils import get_column_letter
import os

MAX_WIDTH = 300
MAX_HEIGHT = 300

def resize(img):
    w, h = img.size
    if w > MAX_WIDTH:
        h = MAX_WIDTH / w * h
        w = MAX_WIDTH

    if h > MAX_HEIGHT:
        w = MAX_HEIGHT / h * w
        h = MAX_HEIGHT

    return img.resize((int(w), int(h)), Image.ANTIALIAS)

# 十进制转十六进制
def int_to_16(num):
    num1 = hex(num).replace('0x', '')
    return num1 if len(num1) > 1 else '0' + num1


def draw_jpg(img_path):
    # 修改图片大小
    img_pic = resize(Image.open(img_path))
    img_name = os.path.basename(img_path)

    print(img_name)
    out_file = './result_' + img_name.split('.')[0] + '.xlsx'
    print(out_file)

    if os.path.exists(out_file):
        os.remove(out_file)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    width, height = img_pic.size
    print(width, height)

    for w in range(1, width + 1):
        for h in range(1, height + 1):
            if img_pic.mode == 'RGB':
                r, g, b = img_pic.getpixel((w - 1, h - 1))
            elif img_pic.mode == 'RGBA':
                r, g, b, a = img_pic.getpixel((w - 1, h - 1))

            hex_rgb = int_to_16(r) + int_to_16(g) + int_to_16(b)

            cell = worksheet.cell(column=w, row=h)

            if h == 1:
                _w = get_column_letter(cell.column) # 把列 idx 转换成字母
                _h = cell.col_idx
                # 调整列宽
                worksheet.column_dimensions[_w].width = 1
            # 调整行高
            worksheet.row_dimensions[h].height = 6

            cell.fill = fills.PatternFill(fill_type="solid", fgColor=hex_rgb)

        print('write in:', w, '  |  all:', width + 1)
    print('saving...')
    workbook.save(out_file)
    print('success!')


if __name__ == '__main__':
    draw_jpg('pro1.jpg')
